import datetime
import traceback
import xlsxwriter
from openpyxl import load_workbook
import threading
from kivy.uix.boxlayout import BoxLayout
from kivy.utils import get_color_from_hex as rgb
from kivy.clock import Clock, mainthread
from kivy.lang import Builder
from kivy.properties import StringProperty, ObjectProperty, BooleanProperty
from kivy.logger import Logger
from screens.base import BaseScreen
from settings import DEVICES, DEVICES_LONG, INTERVAL
from utils.db import get_sensor_data, drop_db, get_raw_sensor_data, group_data
from utils.usb import get_connected_usb_device, save_file_to_usb_drive
from utils.usb import get_file_list_from_usb_drive, load_file_from_usb_drive
from widgets.dialog import YesNoDialog
from widgets.file_list_dialog import FileListDialog, WorksheetListDialog
import numpy as np
from widgets.graph import Graph, SmoothLinePlot
from widgets.label import H4
from widgets.snackbar import Snackbar


Builder.load_file('screens/chart_screen.kv')


graph_theme = {
    'label_options': {
        'color': rgb('444444'),             # color of tick labels and titles
        'bold': True
    },
    'tick_color': rgb('A0A0A0'),            # ticks and grid
    'border_color': rgb('808080')}          # border drawn around each graph


class ChartScreen(BaseScreen):

    colors = {
        'LF': rgb('039BE5'),
        'RF': rgb('F4511E'),
        'LR': rgb('43A047'),
        'RR': rgb('8E24AA'),
        'A': rgb('424242'),
        'B': rgb('F06292'),
        'C': rgb('26A69A'),
    }
    colors_disabled = {
        'LF': rgb('81D4FA'),
        'RF': rgb('FFAB91'),
        'LR': rgb('A5D6A7'),
        'RR': rgb('CE93D8'),
        'A': rgb('BDBDBD'),
        'B': rgb('F8BBD0'),
        'C': rgb('80CBC4'),
    }

    wheels_enabled = {
        "LF": BooleanProperty(True),
        "RF": BooleanProperty(False),
        "LR": BooleanProperty(False),
        "RR": BooleanProperty(False),
    }
    wheel_ids = {
        "LF": "wheel_lf",
        "RF": "wheel_rf",
        "LR": "wheel_lr",
        "RR": "wheel_rr",
    }

    comparisons = {
        "A": None,
        "B": None,
        "C": None
    }

    comparisons_enabled = {
        "A": BooleanProperty(False),
        "B": BooleanProperty(False),
        "C": BooleanProperty(False)
    }

    comparison_ids = {
        "A": "comparison_a",
        "B": "comparison_b",
        "C": "comparison_c"
    }
    comparison_clear_ids = {
        "A": "comparison_a_clear",
        "B": "comparison_b_clear",
        "C": "comparison_c_clear"
    }

    cur_wheel = StringProperty('LF')
    cur_dev_type = StringProperty('weight_wheel')

    _clk_chart = ObjectProperty(allownone=True)
    def load_workbook(self, temp):
        pass

    def on_enter(self, *args):
        super(ChartScreen, self).on_enter(*args)
        self.ids.btn_play.icon = 'pause' if self.app.is_recording else 'play'
        if self.app.is_recording:
            self._clk_chart = Clock.schedule_interval(lambda dt: self._update_chart(), INTERVAL)

        self._update_wheel_buttons()
        self._update_comparison_buttons()
        self._update_chart()

    def on_pre_leave(self, *args):
        super(ChartScreen, self).on_enter(*args)
        if self._clk_chart:
            self._clk_chart.cancel()
            self._clk_chart = None

    def _update_wheel_buttons(self):
        for wheel in self.wheels_enabled:
            self.wheels_enabled[wheel] = True if self.ids[self.wheel_ids[wheel]].state == 'down' else False

            self.ids[self.wheel_ids[wheel]].background_color = \
                self.colors[wheel] if self.ids[self.wheel_ids[wheel]].state == 'down' else self.colors_disabled[wheel]

    def _update_comparison_buttons(self):
        for comp in self.comparisons_enabled:
            self.comparisons_enabled[comp] = True if self.ids[self.comparison_ids[comp]].state == 'down' else False
            self.ids[self.comparison_ids[comp]].background_color = \
                self.colors[comp] if self.ids[self.comparison_ids[comp]].state == 'down' else self.colors_disabled[comp]
            self.ids[self.comparison_clear_ids[comp]].background_color = \
                self.colors[comp] if self.comparisons[comp] is not None else self.colors_disabled[comp]

    def _reset_comparison_button(self, comp):
        self.ids[self.comparison_ids[comp]].state = 'down' if self.comparisons_enabled[comp] else 'normal'

    def on_wheel_changed(self):
        self._update_wheel_buttons()
        self._update_chart()

    def on_dev_type_changed(self):
        self.cur_dev_type = 'weight_wheel' if self.ids.dev_type.get_value() == 'Wheel Weight' else 'weight_bs'
        self._update_chart()

    def on_comparison_changed(self, comparison):
        if self.comparisons[comparison] is None and self.ids[self.comparison_ids[comparison]].state == 'down':
            self.launch_import_dialog(comparison)
            self._reset_comparison_button(comparison)
        else:
            self._update_comparison_buttons()
            self._update_chart()

    def on_comparison_delete(self, comparison):

        self.comparisons_enabled[comparison] = False
        self._reset_comparison_button(comparison)

        # reset this AFTER updating the button to avoid re-triggering the import dialog
        if self.comparisons[comparison] is not None:
            self.comparisons[comparison] = None

        self._update_comparison_buttons()
        self._update_chart()

    def _update_chart(self):
        threading.Thread(target=self._draw_chart()).start()

    def _generate_chart(self):
        charts = {}

        # pull chart data for each selected tire
        for wheel in self.wheels_enabled:
            if self.wheels_enabled[wheel]:
                charts[wheel] = get_sensor_data(wheel)

        # pull chart data for enabled overlays
        for comp in self.comparisons_enabled:
            if self.comparisons_enabled[comp] and self.comparisons[comp] is not None:
                charts[comp] = group_data(self.comparisons[comp])

        # prepare some global values to be updated while creating plot values
        global_max_x = 0
        global_max_y = 0
        global_min_x = float('inf')
        global_min_y = float('inf')

        plots = []

        # process data for each selected tire
        for tire, data in charts.items():
            x_values = []
            y_values = []

            # show only last 50 elements
            if len(data) > 50:
                data = data[-50:-1]

            # sort data by potentiometer reading
            data = sorted(data, key=lambda v: v['potentiometer'])

            # separate each row of data into an x/y value
            for d in data:
                x_values.append(d['potentiometer'])
                y_values.append(d[self.cur_dev_type])

            # don't bother generating a plot layout if no data is present (or only one row)
            len_x = len(x_values)
            if len_x <= 1:
                continue

            # cache min/max/length for the x_values list as they will be used often
            min_x = min(x_values)
            max_x = max(x_values)

            # generate an evenly-spaced list using the min and max x values
            x_new = np.arange(min_x, max_x, (max_x - min_x) / float(len_x))

            # constrain the generated list to the correct min/max values (is this necessary?)
            x_new = [min(x, float(max_x)) for x in x_new]
            x_new = [max(x, float(min_x)) for x in x_new]

            # interpolate the y values for this generated x list
            y_new = np.interp(x_new, x_values, y_values)

            max_x_new = max(x_new)
            min_x_new = min(x_new)
            max_y_new = max(y_new)
            min_y_new = min(y_new)

            # record global min/max/length across all enabled plots, used to create the graph limits later
            if max_x_new > global_max_x:
                global_max_x = max_x_new
            if min_x_new < global_min_x:
                global_min_x = min_x_new
            if max_y_new > global_max_y:
                global_max_y = max_y_new
            if min_y_new < global_min_y:
                global_min_y = min_y_new

            plot = SmoothLinePlot(color=self.colors[tire])
            plot.points = zip(x_new, y_new)
            plots.append(plot)

        # return an empty layout if no plots were generated
        if len(plots) < 1:
            return None
            
        # create a layout to hold the graph
        w, h = self.ids.box.size
        p = self.ids.box.padding[0]
        layout = BoxLayout(orientation='vertical', opacity=0, size_hint=(None, None), size=(w - 2 * p, h - 2 * p))

        # generate graph
        graph = Graph(
                xlabel='Travel',
                ylabel='Wheel Weight' if self.cur_dev_type == 'weight_wheel' else 'Bump Load',

                x_ticks_major=(global_max_x - global_min_x) / 5.,
                x_ticks_minor=5,
                y_ticks_major=(global_max_y - global_min_y) / 5.,
                y_ticks_minor=5,

                y_grid_label=True,
                x_grid_label=True,
                padding=5,
                xlog=False,
                ylog=False,
                x_grid=True,
                y_grid=True,
                x_unit='in',

                xmin=float(global_min_x),
                xmax=float(global_max_x if global_max_x != global_min_x else global_min_x + 1),
                ymin=float(global_min_y),  # chart should only be generated when there are values in y
                ymax=float(global_max_y if global_max_y != global_min_y else global_min_y + 1),

                _with_stencilbuffer=False,
                font_size='12sp',
                precision='%.1f',
                **graph_theme)

        # populate graph with plots
        for plot in plots:
            graph.add_plot(plot)

        layout.add_widget(graph)

        return layout
            
    def _draw_chart(self):

        try:
            layout = self._generate_chart()

            Clock.schedule_once(lambda dt: self._add_chart_widget(layout))

        except Exception as e:
            Logger.exception('Failed to draw the chart - {}'.format(e))
            self.app.save_exception(traceback.format_exc(limit=20))
            self.app.switch_screen('error_screen')
                
    @mainthread
    def _add_chart_widget(self, layout):
        self.ids.box.clear_widgets()
        if layout:
            self.ids.box.add_widget(layout)
            Clock.schedule_once(lambda dt: self.show_layout(layout), .1)
        else:
            self.ids.box.add_widget(H4(text='Chart Data Not Available', font_style='Display1', halign='center'))

    @staticmethod
    def show_layout(layout):
        layout.opacity = 1

    def on_btn_download(self):
        dev_path = get_connected_usb_device()
        if dev_path is None:
            Snackbar(text="   Cannot find any USB Drive!   ", background_color=(.8, 0, .3, .5)).show()
            return
        threading.Thread(target=self._save_data_to_usb, args=(dev_path, )).start()

    def launch_import_dialog(self, comparison):
        dev_path = get_connected_usb_device()
        if dev_path is None:
            Snackbar(text="   Cannot find any USB Drive!   ", background_color=(.8, 0, .3, .5)).show()
            return

        file_list = get_file_list_from_usb_drive(dev_path)
        if file_list is None or len(file_list) == 0:
            Snackbar(text="   Cannot find any valid files!   ", background_color=(.8, 0, .3, .5)).show()
            return

        dlg = FileListDialog(file_list, comparison)
        dlg.bind(on_confirm=self.on_confirm_file_choice)
        dlg.open()

    def on_confirm_file_choice(self, instance, filename, comparison):
        # open worksheet choice with sheets in file
        # then store as one of the 3 available overlays?

        dev_path = get_connected_usb_device()
        if dev_path is None:
            Snackbar(text="   Cannot find any USB Drive!   ", background_color=(.8, 0, .3, .5)).show()
            return

        temp_file = load_file_from_usb_drive(dev_path, filename)
        if temp_file is None:
            Snackbar(text="   Cannot load file from USB Drive!   ", background_color=(.8, 0, .3, .5)).show()
            return

        workbook = load_workbook(temp_file)

        dlg = WorksheetListDialog(temp_file, comparison, workbook.get_sheet_names())
        dlg.bind(on_confirm=self.on_confirm_worksheet_choice)
        dlg.open()

    def on_confirm_worksheet_choice(self, instance, temp_file, worksheet, comparison):
        self._read_data_from_usb(temp_file, worksheet, comparison)

    def _read_data_from_usb(self, file_path, wheel, comparison):
        Snackbar(text="Loading...").show()
        workbook = load_workbook(file_path)

        try:
            sheet = workbook.get_sheet_by_name(wheel)
        except KeyError as e:
            Clock.schedule_once(lambda dt: Snackbar(text="Cannot load the selected wheel from this file. Please try a different wheel or file.", background_color=(.8, 0, .3, .5), duration=4).show())
            return

        row_start = 2
        data = []
        for row in range(row_start, sheet.max_row+1):
            data.append({
                'ts': sheet.cell(row=row, column=1).value,  # using timestamp to sort the data later
                'weight_wheel': float(sheet.cell(row=row, column=2).value),
                'weight_bs': float(sheet.cell(row=row, column=3).value),
                'potentiometer': float(sheet.cell(row=row, column=4).value)
            })

        # sort sheet by timestamp (just in case it's out of order for some reason)
        data = sorted(data, key=lambda r: r['ts'])

        self.comparisons[comparison] = data
        self.comparisons_enabled[comparison] = True
        self._reset_comparison_button(comparison)

        Clock.schedule_once(lambda dt: Snackbar(text="Loaded chart data from file").show())
        Clock.schedule_once(lambda dt: self._update_comparison_buttons())
        Clock.schedule_once(lambda dt: self._update_chart())

    def _save_data_to_usb(self, dev_path):
        # Create a xlsx file and export data.
        file_name = 'WSD_Export_{}.xlsx'.format(datetime.datetime.now().isoformat().replace(':', '-').split('.')[0])
        file_path = '/tmp/' + file_name
        workbook = xlsxwriter.Workbook(file_path)
        for wheel in DEVICES:
            data = get_raw_sensor_data(wheel)
            # sort by timestamp
            #data = sorted(raw_data, key=lambda v: v['ts'])
            worksheet = workbook.add_worksheet(wheel)
            worksheet.write(0, 0, 'DateTime')
            worksheet.write(0, 1, 'Wheel Weight')
            worksheet.write(0, 2, 'Bump Load')
            worksheet.write(0, 3, 'Travel')
            for i, d in enumerate(data):
                worksheet.write(i + 1, 0, datetime.datetime.fromtimestamp(d['ts']).isoformat())
                worksheet.write(i + 1, 1, d['weight_wheel'])
                worksheet.write(i + 1, 2, d['weight_bs'])
                worksheet.write(i + 1, 3, d['potentiometer'])
        workbook.close()

        result = save_file_to_usb_drive(dev_path, file_path)

        Clock.schedule_once(lambda dt: self._show_result(result, file_name))

    @staticmethod
    def _show_result(result, file_name):
        if result:
            Clock.schedule_once(lambda dt: Snackbar(text="Successfully exported - {}".format(file_name), ).show())
        else:
            Clock.schedule_once(lambda dt: Snackbar(text="Failed to export", background_color=(.8, 0, .3, .5)).show())

    def on_btn_clear(self):
        dlg = YesNoDialog(message='Are you sure to clear all data?', title='Wireless Scale Display')
        dlg.bind(on_confirm=self.on_confirm_clear)
        dlg.open()

    def on_confirm_clear(self, *args):
        self.ids.box.clear_widgets()
        threading.Thread(target=drop_db).start()
        Snackbar(text="Cleared all data").show()
        args[0].dismiss()

    def on_btn_play(self, btn):
        if btn.icon == 'pause':
            btn.icon = 'play'
            self.app.stop_recording()
            if self._clk_chart:
                self._clk_chart.cancel()
                self._clk_chart = None
            Snackbar(text='Stopped Recording').show()
        else:
            btn.icon = 'pause'
            self.app.start_recording()
            self._clk_chart = Clock.schedule_interval(lambda dt: self._update_chart(), INTERVAL)
            self._update_chart()
            Snackbar(text='Started Recording').show()
